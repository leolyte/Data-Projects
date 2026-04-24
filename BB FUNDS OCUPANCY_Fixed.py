# -*- coding: utf-8 -*-
"""
Created on Tue Feb 17 14:05:25 2026
@author: LCHIBUIKE

Build TWO sheets:

1) Master_Comparison:
   Samples | Wilmut 0,5 | Wilmut 1.4mL | 9x9 | Gradetes
   + Total/ponderation columns
   + TOTAL row
   + Percentage displayed as % (Excel formatting)

2) CRIOBOX_Nitrogen_10x10:
   Samples | CRIOBOX Nitrogen 10x10

"""

import os
import glob
import unicodedata
import pandas as pd

# =========================
# CONFIG
# =========================
# >>> UPDATE THIS LINE BEFORE EACH RUN <<<
# LOOKUP_PATH: exact filename of the Box Type lookup file exported from Access.
#   Example: "BOX_TYPE_16022026.xlsx"
# 
# Place all _SAMPLES_ files from Noray Banks LIMS in the same folder as this script.
# The script will find them automatically using the pattern below — no path needed.
SAMPLES_GLOB = "*_SAMPLES_*.xlsx"
LOOKUP_PATH  = "BOX_TYPE_16022026.xlsx"
OUT_XLSX     = "Master_Comparison_All_Files.xlsx"

MASTER_SHEET = "Master_Comparison"
CRIO_SHEET   = "CRIOBOX_Nitrogen_10x10"

SEARCH_ROWS_FOR_HEADER = 200

# Columns in the Master sheet (CRIO removed)
TARGET_COLS = ["Samples", "Wilmut 0,5", "Wilmut 1.4mL", "9x9", "Gradetes"]


# =========================
# HELPERS
# =========================
def strip_accents(s):
    if pd.isna(s):
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(s))
        if not unicodedata.combining(c)
    )

def normalize_code(s):
    if pd.isna(s):
        return ""
    s = str(s).strip()
    return " ".join(s.split()).upper()

def normalize_text(s):
    if pd.isna(s):
        return ""
    return str(s).strip().upper()

def detect_header_and_load(path, expected_col="Box name", search_rows=200):
    raw = pd.read_excel(path, header=None, dtype=str, engine="openpyxl")
    header_idx = None
    for i in range(min(search_rows, raw.shape[0])):
        row_vals = raw.iloc[i].astype(str).str.strip().str.lower().tolist()
        if expected_col.lower() in row_vals:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"[{os.path.basename(path)}] Could not find header row with '{expected_col}'.")
    header = raw.iloc[header_idx].astype(str).str.strip().tolist()
    df = raw.iloc[header_idx+1:].copy()
    df.columns = header
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df = df.dropna(how="all")
    df.columns = [c.strip() for c in df.columns]
    return df.reset_index(drop=True)

def extract_file_id(filename: str) -> str:
    up = filename.upper()
    if "_SAMPLES" in up:
        return up.split("_SAMPLES")[0]
    return os.path.splitext(filename)[0].upper()


# =========================
# LOAD LOOKUP
# =========================
df_lookup = pd.read_excel(LOOKUP_PATH, dtype=str, engine="openpyxl")
df_lookup.columns = df_lookup.columns.str.strip()

df_lookup["__key"] = df_lookup["Labware"].map(normalize_code).astype(str)
df_lookup_clean = df_lookup.dropna(subset=["UBIC_TIPUS"])

df_lookup_dedup = (
    df_lookup_clean
    .sort_values("__key")
    .drop_duplicates("__key", keep="last")[["__key", "UBIC_TIPUS"]]
)

df_lookup_dedup["__key"] = df_lookup_dedup["__key"].astype(str)


# =========================
# PROCESS ALL FILES
# =========================
rows = []       # For Master sheet
crio_rows = []  # For CRIOBOX Nitrogen 10x10 sheet
unknowns = set()

for path in glob.glob(SAMPLES_GLOB):
    filename = os.path.basename(path)
    print(f"Processing: {filename}")

    df = detect_header_and_load(path, search_rows=SEARCH_ROWS_FOR_HEADER)

    # Filter TRÀNSIT / TRANSIT / Exhausted
    df["__C"]  = df["Container"].map(normalize_text)
    df["__CA"] = df["__C"].map(strip_accents)
    df["__S"]  = df["Sample status"].map(normalize_text)

    mask = (
        (df["__C"] == "TRÀNSIT") |
        (df["__CA"] == "TRANSIT") |
        (df["__S"] == "EXHAUSTED")
    )
    df = df.loc[~mask].copy()
    df.drop(columns=["__C","__CA","__S"], inplace=True, errors="ignore")

    # Merge with lookup
    df["__key"] = df["Box name"].map(normalize_code).astype(str)
    merged = df.merge(df_lookup_dedup, how="left", on="__key")

    # Ensure both Box name & UBIC_TIPUS are strings before combining
    merged["UBIC_TIPUS"] = merged["UBIC_TIPUS"].astype("string")
    merged["Box name"]   = merged["Box name"].astype("string")

    merged["Box name"] = merged["UBIC_TIPUS"].fillna(merged["Box name"])

    # Ensure Box name is always string dtype before assigning text labels
    merged["Box name"] = merged["Box name"].astype("string")

    # -------------------------
    # Apply renaming rules
    # -------------------------
    merged["__box_norm"] = (
        merged["Box name"]
        .astype(str)
        .str.strip()
        .str.upper()
        .map(strip_accents)
    )

    # Wilmut 0.5
    p05 = [
        "CALAIX AMPLE PER MICROPLAQUES (-80ºC)",
        "CALAIX ESTRET PER MICROPLAQUES (-80ºC)",
        "CALAIX SOST MICROPLAQUES (-80ºC)"
    ]
    merged.loc[
        merged["__box_norm"].isin([strip_accents(p.upper()) for p in p05]),
        "Box name"
    ] = "Wilmut 0,5"

    # Wilmut 1.4mL
    p14 = [
        "CALAIX PER MICROPLAQUES GRANS 1,4ML (-80ºC)",
        "CALAIX SOST MICROPLAQUES GRANS (-80ºC)"
    ]
    merged.loc[
        merged["__box_norm"].isin([strip_accents(p.upper()) for p in p14]),
        "Box name"
    ] = "Wilmut 1.4 mL"

    # 9x9
    p9 = [
        "CALAIX CAIXES ALTES 9X9 (-80ºC)",
        "CALAIX PER CAIXES 9X9 (-80ºC)",
        "CALAIX SOST CAIXES 9X9 (-80ºC)"
    ]
    merged.loc[
        merged["__box_norm"].isin([strip_accents(p.upper()) for p in p9]),
        "Box name"
    ] = "9x9"

    # Gradetes
    pgrad = [
        "ESTANT PER GRADETES 10X4",
        "GRADETES TUBS 5X12",
        "GRADETES TUBS 6 POSICIONES",
        "GRADETES TUBS 4 POSICIONES"
    ]
    merged.loc[
        merged["__box_norm"].isin([strip_accents(p.upper()) for p in pgrad]),
        "Box name"
    ] = "Gradetes"

    # Count final box names
    counts = (
        merged["Box name"]
        .astype("string")
        .value_counts(dropna=False)
        .astype(int)
        .to_dict()
    )

    # Track unknowns
    allowed = {"Wilmut 0,5", "Wilmut 1.4 mL", "9x9", "Gradetes", "CRIOBOX Nitrogen 10x10"}
    for b in counts:
        if b not in allowed:
            unknowns.add(b)

    file_id = extract_file_id(filename)

    # MASTER ROW (CRIO removed)
    rows.append({
        "Samples": file_id,
        "Wilmut 0,5": counts.get("Wilmut 0,5", 0),
        "Wilmut 1.4mL": counts.get("Wilmut 1.4 mL", 0),  # note: source has space before mL
        "9x9": counts.get("9x9", 0),
        "Gradetes": counts.get("Gradetes", 0)
    })

    # CRIOBOX SHEET ROW
    crio_rows.append({
        "Samples": file_id,
        "CRIOBOX Nitrogen 10x10": counts.get("CRIOBOX Nitrogen 10x10", 0)
    })


# =========================
# BUILD DATAFRAMES
# =========================
master = pd.DataFrame(rows)
crio   = pd.DataFrame(crio_rows)

# Guarantee master columns/order
for col in TARGET_COLS:
    if col not in master.columns:
        master[col] = 0

master = master[TARGET_COLS].sort_values("Samples").reset_index(drop=True)

# ---- Add capacity-based totals (float, 2 decimals)
master["Total Wilmut 0,5"]   = (master["Wilmut 0,5"]   / 96).astype(float).round(2)
master["Total Wilmut 1.4mL"] = (master["Wilmut 1.4mL"] / 96).astype(float).round(2)
master["Total 9x9"]          = (master["9x9"]          / 81).astype(float).round(2)
master["Total Gradetes"]     = (master["Gradetes"]     / 60).astype(float).round(2)

# ---- Add PONDERADES (weighted values from TOTAL columns)
master["PONDERADES Wilmut 0,5"]   = (master["Total Wilmut 0,5"]   * 1).round(2)
master["PONDERADES Wilmut 1.4mL"] = (master["Total Wilmut 1.4mL"] * 2).round(2)
master["PONDERADES 9x9"]          = (master["Total 9x9"]          * 2.55).round(2)
master["PONDERADES Gradetes"]     = (master["Total Gradetes"]     * 10.2).round(2)

# ---- Reorder final Master columns
master = master[
    [
        "Samples",
        "Wilmut 0,5", "Total Wilmut 0,5", "PONDERADES Wilmut 0,5",
        "Wilmut 1.4mL", "Total Wilmut 1.4mL", "PONDERADES Wilmut 1.4mL",
        "9x9", "Total 9x9", "PONDERADES 9x9",
        "Gradetes", "Total Gradetes", "PONDERADES Gradetes"
    ]
]

# ============================================
# ADD TOTAL PONDERATION + PERCENTAGE COLUMNS
# ============================================
# 1) TOTAL PONDERATION (sum of ponderations per row)
master["TOTAL PONDERATION"] = (
    master["PONDERADES Wilmut 0,5"] +
    master["PONDERADES Wilmut 1.4mL"] +
    master["PONDERADES 9x9"] +
    master["PONDERADES Gradetes"]
).round(2)

# 2) PERCENTAGE PONDERATION (store as decimal; Excel will render as %)
total_ponderation_sum = master["TOTAL PONDERATION"].sum()
master["PERCENTAGE PONDERATION"] = (
    master["TOTAL PONDERATION"] / total_ponderation_sum
).round(4)

# 3) Append TOTAL row
total_row = {"Samples": "TOTAL"}
for col in master.columns:
    if col != "Samples":
        total_row[col] = round(master[col].sum(), 2)
# Ensure percentage is exactly 100%
total_row["PERCENTAGE PONDERATION"] = 1.00  # decimal => 100% in Excel
master = pd.concat([master, pd.DataFrame([total_row])], ignore_index=True)

# ---- CRIO sheet unchanged
crio = crio[["Samples", "CRIOBOX Nitrogen 10x10"]].sort_values("Samples").reset_index(drop=True)

# =========================
# SAVE BOTH SHEETS + FORMAT
# =========================
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    master.to_excel(writer, sheet_name=MASTER_SHEET, index=False)
    crio.to_excel(writer, sheet_name=CRIO_SHEET, index=False)

    unknowns_clean = {str(x) for x in unknowns if x is not pd.NA and not pd.isna(x)}
    unknown_df = pd.DataFrame(sorted(unknowns_clean), columns=["Unknown Box Names"])
    unknown_df.to_excel(writer, sheet_name="Unknown_Categories", index=False)


    # ===============================
    # APPLY IDIBAPS BLUE FORMATTING
    # ===============================
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = writer.book

    # --- IDIBAPS Colours ---
    blue = "0055A4"       # Dark IDIBAPS blue
    light_blue = "66A3D2"  # Accent blue
    gray = "F2F2F2"        # Soft banding gray
    white = "FFFFFF"

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    def format_sheet(ws, percent_header="PERCENTAGE PONDERATION", total_label="TOTAL"):
        max_row = ws.max_row
        max_col = ws.max_column

        # --- Header formatting ---
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # --- Zebra banding (exclude total row) ---
        for row in range(2, max_row):
            fill_color = white if row % 2 == 0 else gray
            for col in range(1, max_col + 1):
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=fill_color)
                c.border = thin_border

        # --- TOTAL row highlight (last row) ---
        for col in range(1, max_col + 1):
            cell = ws.cell(row=max_row, column=col)
            cell.fill = PatternFill("solid", fgColor=light_blue)
            cell.font = Font(bold=True)
            cell.border = thin_border
            # Keep numbers centered in total row
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Format % column if present ---
        # Find the column index by header text
        percent_col_idx = None
        for col in range(1, max_col + 1):
            if str(ws.cell(row=1, column=col).value).strip().upper() == percent_header.upper():
                percent_col_idx = col
                break
        if percent_col_idx is not None:
            for row in range(2, max_row + 1):
                pct_cell = ws.cell(row=row, column=percent_col_idx)
                pct_cell.number_format = "0.00%"

        # --- Autofit column width ---
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[col_letter]:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 3, 60)

        # --- Freeze header row ---
        ws.freeze_panes = "A2"

        # --- Add autofilter over used range ---
        ws.auto_filter.ref = ws.dimensions

    # Apply formatting to each sheet
    format_sheet(wb[MASTER_SHEET])   # has percentage column
    format_sheet(wb[CRIO_SHEET])     # does not have percentage; function skips gracefully


import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap

# --------------------------
# Load and prepare data
# --------------------------

df = master.copy()

# Rename columns
df.rename(columns={df.columns[0]: "Cohort"}, inplace=True)
df.rename(columns={
    df.columns[13]: "TOTAL_PONDERATION",
    df.columns[14]: "PERCENTAGE_PONDERATION"
}, inplace=True)

# Remove TOTAL row
df = df[df["Cohort"].astype(str).str.upper() != "TOTAL"]

# Convert numeric
df["TOTAL_PONDERATION"] = pd.to_numeric(df["TOTAL_PONDERATION"], errors="coerce").fillna(0)
df["PERCENTAGE_PONDERATION"] = pd.to_numeric(df["PERCENTAGE_PONDERATION"], errors="coerce").fillna(0)

# Sort by total ponderation
df_plot = df.sort_values(by="TOTAL_PONDERATION", ascending=False)

# --------------------------
# Colour scale
# --------------------------
pct_norm = (df_plot["PERCENTAGE_PONDERATION"] - df_plot["PERCENTAGE_PONDERATION"].min()) / (
    df_plot["PERCENTAGE_PONDERATION"].max() - df_plot["PERCENTAGE_PONDERATION"].min()
)

# Compressed blue gradient (no faint bars)
colors_blue = LinearSegmentedColormap.from_list(
    "idibaps_blue_compressed",
    ["#7FAFEA", "#3F7DDC", "#004A99"]
)
colors = colors_blue(pct_norm)

# --------------------------
# Plotting
# --------------------------
plt.figure(figsize=(18, 9))   # bigger plotting area without changing bars

bars = plt.barh(df_plot["Cohort"], df_plot["TOTAL_PONDERATION"], color=colors)

max_width = df_plot["TOTAL_PONDERATION"].max()
padding = max_width * 0.03   # keep labels nicely outside

# Label each bar
for i, bar in enumerate(bars):
    width = bar.get_width()
    pct_val = df_plot["PERCENTAGE_PONDERATION"].iloc[i] * 100
    
    plt.text(
        width + padding,
        bar.get_y() + bar.get_height()/2,
        f"{width:.2f} ({pct_val:.2f}%)",
        va="center", fontsize=13
    )

# --------------------------
# Ensure labels fit inside the plotting area
# --------------------------
plt.xlim(0, max_width * 1.20)   # <-- this ensures ALL labels stay inside
plt.margins(x=0.02)

# --------------------------
# Formatting
# --------------------------
plt.gca().invert_yaxis()

plt.xlabel("TOTAL PONDERATION", fontsize=14)
plt.ylabel("Cohort", fontsize=14)
plt.title("Ranked TOTAL PONDERATION — Compressed Blue Magnitude Scale", fontsize=18)

# Expand internal figure margins
plt.subplots_adjust(left=0.22, right=0.95, top=0.94, bottom=0.08)

# --------------------------
# Save high‑res PNG
# --------------------------
plt.savefig("TOTAL_PONDERATION_RANKED.png", dpi=300, bbox_inches='tight')
plt.show()



print(f"✔ Saved: {OUT_XLSX}")
if unknowns:
    print("⚠ Unknown categories found:")
    for u in sorted([str(x) for x in unknowns]):
        print("  -", u)
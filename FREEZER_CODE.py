# -*- coding: utf-8 -*-
"""
Created on Tue Feb 10 11:57:46 2026

@author: LCHIBUIKE
"""
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


# --------------------------- CONFIG -----------------------------------------
# >>> UPDATE THIS LINE BEFORE EACH RUN <<<
in_path  = Path("FREEZERS_20260205 (1).xlsx")
out_path = Path("FREEZERS_summary.xlsx")

loc_type_col = "tipus ubic"
loc_col      = "UBICACIO"
occ_col      = "CAPACITAT"
free_col     = "LLIURES"

# ----------------------- 1) Load Excel --------------------------------------
df = pd.read_excel(in_path, sheet_name=0, engine="openpyxl")

# ----------------------- 2) Normalize 'tipus ubic' ---------------------------
r1 = [
    "Calaix ample per microplaques (-80ºC)",
    "Calaix estret per microplaques (-80ºC)",
    "Calaix SOST microplaques (-80ºC)"
]
r2 = [
    "Calaix per microplaques grans 1,4ml (-80ºC)",
    "Calaix SOST microplaques grans (-80ºC)"
]
r3 = [
    "Calaix CAIXES ALTES 9x9 (-80ºC)",
    "Calaix per CAIXES 9x9 (-80ºC)",
    "Calaix SOST CAIXES 9X9 (-80ºC)"
]
r4 = [
    "Estant per Gradetes 10x4",
    "Gradetes tubs 4 posiciones",
    "Gradetes tubs 5x12"
]
r5 = [
    "Columna per 4 CAIXES 9x9 BTT",
    "Columna per 5 CAIXES 9x9 BTT"
]
r6 = [
    "Gradetes tubs 6 posiciones"
]

replacement_map = {
    **{k: "Willmut 0,5"    for k in r1},
    **{k: "Wilmut 1.4 mL"  for k in r2},
    **{k: "9x9"            for k in r3},
    **{k: "Gradetes"       for k in r4},
    **{k: "9x9BTT"         for k in r5},
    **{k: "Gradetes"       for k in r6},
}

if loc_type_col in df.columns:
    df[loc_type_col] = df[loc_type_col].replace(replacement_map)

# ----------------------- 3) Clean UBICACIO ----------------------------------
if loc_col in df.columns:
    df[loc_col] = df[loc_col].astype(str).str.split('_').str[0]

df[loc_col] = df[loc_col].str.replace(r"^LOG1$", "LOG01", regex=True)

# ----------------------- 4) Drop unused columns ------------------------------
cols_to_drop = [
    'CuentaDeLABWARE', 'ID_UBICACIO', 'ubic', 'TOTAL', 'PMAX', 'PDISP',
    'percentatge_disp', 'percentatge_ocup', 'recompte'
]
df = df.drop(columns=[c for c in cols_to_drop if c in df.columns])

# ----------------------- 5) Remove unwanted UBICACIO values ------------------
values_to_remove = {
    'Mostres reubicades a microplaca nova.', 'NE2', 'NE3', 'NE1',
    'Nevera 4ºC Cultius', 'Nevera 4ºC aDNA',
    'PENDENT INVENTARI', 'PRE01', 'PRE02', 'PRE03',
    'Congelador -20ºC aDNA', 'Congelador -20ºC', 'PRE04',
    'ELIMINADA', 'EN TRÀNSIT PER CESSIÓ', 'IP', 'LOG05',
    'Primer Envío MATER', 'RECUPERADA PER IP',
    'REUBICACIO MOSTRES A ALTRA CAIXA', 'TUM3', 'TUM4',
    'SOST10', 'Noelia', 'TRÀNSIT','LOG01','DNM'
}
df = df[~df[loc_col].isin(values_to_remove)].copy()

# ----------------------- 6) Drop rows by tipus ubic patterns -----------------
drop_terms = ["Calaix Alt", "Envíos MII", "Envios MII", "TRÀNSIT", "TRANSIT"]
df = df[~df[loc_type_col].astype(str).str.contains("|".join(drop_terms), case=False, na=False)].copy()

# ----------------------- 7) Ensure numeric + basic cleaning ------------------
df[occ_col]  = pd.to_numeric(df[occ_col], errors="coerce")
df[free_col] = pd.to_numeric(df[free_col], errors="coerce")
df_clean = df.dropna(subset=[loc_type_col, loc_col]).copy()

# ----------------------- 8) Weights & ponderació -----------------------------
def get_weight(t: str) -> float:
    t = str(t).upper()
    if "9X9" in t or "BTT" in t:
        return 2.55
    if "GRAD" in t:
        return 10.2
    if "1.4" in t:
        return 2.0
    if "0,5" in t:
        return 1.0
    if "ALT" in t:
        return 1.0
    return 0.0

df_clean['WEIGHT'] = df_clean[loc_type_col].apply(get_weight)
df_clean['PONDERACIÓ_TOTAL'] = df_clean[occ_col]  * df_clean['WEIGHT']
df_clean['PONDERACIÓ_FREE']  = df_clean[free_col] * df_clean['WEIGHT']

# ----------------------- 9) Nitrogen extraction & removal --------------------
nitrogen_prefixes = ('NL1', 'NL2', 'NL3', 'NL4', 'NL5')

df_nitrogen = df_clean[df_clean[loc_col].astype(str).str.startswith(nitrogen_prefixes)].copy()
df_clean    = df_clean[~df_clean[loc_col].astype(str).str.startswith(nitrogen_prefixes)].copy()

nitrogen_grouped = (
    df_nitrogen.groupby(loc_col, as_index=False)[
        [occ_col, free_col, 'PONDERACIÓ_TOTAL', 'PONDERACIÓ_FREE']
    ].sum()
)

# ----------------------- 10) Build main outputs (no nitrogen) ----------------
grouped = (
    df_clean.groupby([loc_col, loc_type_col], as_index=False)[
        [occ_col, free_col, 'PONDERACIÓ_TOTAL', 'PONDERACIÓ_FREE']
    ].sum()
)

by_type = (
    df_clean.groupby(loc_type_col, as_index=False)[
        [occ_col, free_col, 'PONDERACIÓ_TOTAL', 'PONDERACIÓ_FREE']
    ].sum()
)

by_location_totals = (
    df_clean.groupby('UBICACIO', as_index=False)[
        ['PONDERACIÓ_TOTAL', 'PONDERACIÓ_FREE']
    ].sum()
)

by_location_totals = by_location_totals.rename(columns={
    'PONDERACIÓ_TOTAL': 'Sum of PONDERACIÓ_TOTAL',
    'PONDERACIÓ_FREE':  'Sum of PONDERACIÓ_FREE'
})

total_free_space = round(by_location_totals["Sum of PONDERACIÓ_FREE"].sum(), 1)
total_space = round(by_location_totals["Sum of PONDERACIÓ_TOTAL"].sum(), 1)

growth_df = pd.DataFrame({
    "Box_Type": ["Willmut 0,5", "Wilmut 1.4 mL", "9x9", "Gradetes"],
    "Quantity": [
        round(total_free_space / 1,    1),
        round(total_free_space / 2,    1),
        round(total_free_space / 2.55, 1),
        round(total_free_space / 10.2, 1)
    ]
})

# ----------------------- 12) Save Excel (with formatting) ----------------------
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    grouped.to_excel(writer, sheet_name="By_Location_and_Type", index=False)
    by_type.to_excel(writer, sheet_name="Totals_by_Type", index=False)
    nitrogen_grouped.to_excel(writer, sheet_name="Nitrogen_Tanks", index=False)
    by_location_totals.to_excel(writer, sheet_name="Totals_Ponderacio_by_UBICACIO", index=False)

    sheet = "Max_Growth_By_BoxType"

    table1 = pd.DataFrame({
        "Metric": ["Total Free Space (Weighted)"],
        "Value": [total_free_space]
    })
    table1.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0, index=False)

    start_row_table2 = len(table1) + 3
    growth_df.to_excel(writer, sheet_name=sheet, startrow=start_row_table2, startcol=0, index=False)

    wb = writer.book

    THIN = Side(style="thin", color="D9D9D9")
    border_all = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    header_fill = PatternFill("solid", fgColor="E9EFF7")
    header_font = Font(bold=True, color="000000")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    def autofit_columns(ws):
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col_idx).value
                if val is None:
                    continue
                val_str = f"{val:,}" if isinstance(val, (int, float)) else str(val)
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(10, min(48, max_len + 2))

    def style_header_row(ws, header_row=1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all

    def apply_number_formats(ws, header_row=1):
        numfmt_map = {
            "CAPACITAT": "#,##0",
            "LLIURES": "#,##0",
            "OCC": "#,##0",
            "TOTAL": "#,##0",
            "FREE": "#,##0",
            "PONDERACIÓ_TOTAL": "#,##0.0",
            "PONDERACIÓ_FREE": "#,##0.0",
            "Sum of PONDERACIÓ_TOTAL": "#,##0.0",
            "Sum of PONDERACIÓ_FREE": "#,##0.0",
            "Value": "#,##0.0",
            "Quantity": "#,##0.0",
        }

        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            fmt = numfmt_map.get(str(header_val), None)
            col_map[col_idx] = fmt

        for row in range(header_row + 1, ws.max_row + 1):
            for col_idx, fmt in col_map.items():
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)) and fmt:
                    cell.number_format = fmt
                    cell.alignment = align_right
                else:
                    if fmt is None:
                        cell.alignment = align_left
                cell.border = border_all

    def add_excel_table(ws, min_row, max_row, min_col, max_col, name, style="TableStyleMedium9"):
        if max_row < min_row or max_col < min_col:
            return
        if max_row - min_row + 1 < 2:
            return

        ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

        wb = ws.parent
        existing_names = set()
        for w in wb.worksheets:
            try:
                existing_names |= set(getattr(w, "tables", {}).keys())
            except Exception:
                tlist = getattr(w, "_tables", [])
                for t in tlist:
                    try:
                        existing_names.add(t.displayName)
                    except Exception:
                        pass

        base = "".join(ch if ch.isalnum() else "_" for ch in str(name)) or "Tbl"
        tname = base
        i = 1
        while tname in existing_names:
            i += 1
            tname = f"{base}_{i}"

        table = Table(displayName=tname, ref=ref)
        style_info = TableStyleInfo(
            name=style,
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
            showColumnStripes=False
        )
        table.tableStyleInfo = style_info
        ws.add_table(table)

    for sname in ["By_Location_and_Type", "Totals_by_Type", "Nitrogen_Tanks", "Totals_Ponderacio_by_UBICACIO"]:
        ws = writer.sheets[sname]
        ws.freeze_panes = "A2"
        style_header_row(ws, header_row=1)
        apply_number_formats(ws, header_row=1)
        add_excel_table(ws, min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column,
                        name=f"Tbl_{sname}")
        autofit_columns(ws)

    ws = writer.sheets[sheet]

    t1_header_row = 1
    t1_rows = len(table1) + 1
    t1_min_col, t1_max_col = 1, 2

    style_header_row(ws, header_row=t1_header_row)
    apply_number_formats(ws, header_row=t1_header_row)
    add_excel_table(ws,
                    min_row=t1_header_row,
                    max_row=t1_rows,
                    min_col=t1_min_col,
                    max_col=t1_max_col,
                    name="Tbl_Max_Free")

    t2_header_row = start_row_table2 + 1
    t2_rows = t2_header_row + len(growth_df)
    for col_idx in range(1, ws.max_column + 1):
        if col_idx <= 2:
            cell = ws.cell(row=t2_header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all

    t2_headers = {ws.cell(row=t2_header_row, column=c).value: c for c in range(1, 3)}
    for r in range(t2_header_row + 1, t2_header_row + 1 + len(growth_df)):
        c_box = t2_headers.get("Box_Type", 1)
        cell = ws.cell(row=r, column=c_box)
        cell.alignment = align_left
        cell.border = border_all

        c_qty = t2_headers.get("Quantity", 2)
        cell = ws.cell(row=r, column=c_qty)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.0"
        cell.alignment = align_right
        cell.border = border_all

    add_excel_table(ws,
                    min_row=t2_header_row,
                    max_row=t2_rows,
                    min_col=1,
                    max_col=2,
                    name="Tbl_Max_Growth")

    ws.freeze_panes = "A2"
    autofit_columns(ws)

print("DONE — Your Excel file is ready with formatted tables!")


# =============================================================================
# MATPLOTLIB CHART — NORMAL (no broken axis)
# =============================================================================
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import StrMethodFormatter

# ---------------------------------------------------------------------
# 1) LOAD AND PREPARE DATA
# ---------------------------------------------------------------------
pivot = by_location_totals.copy()

for c in pivot.columns:
    if "UBICACIO" in str(c):
        pivot.rename(columns={c: "Location"}, inplace=True)
    if "TOTAL" in str(c).upper():
        pivot.rename(columns={c: "TOTAL"}, inplace=True)
    if "FREE" in str(c).upper():
        pivot.rename(columns={c: "FREE"}, inplace=True)

df = pivot[["Location", "TOTAL", "FREE"]].dropna(subset=["Location"]).copy()
df["TOTAL"] = pd.to_numeric(df["TOTAL"], errors="coerce").fillna(0)
df["FREE"] = pd.to_numeric(df["FREE"], errors="coerce").fillna(0).clip(lower=0)
df["OCC"] = (df["TOTAL"] - df["FREE"]).clip(lower=0)
df.sort_values(by="TOTAL", ascending=False, inplace=True)

# Percentages
df["PCT_FREE"] = np.where(df["TOTAL"] > 0, df["FREE"] / df["TOTAL"] * 100, 0)
df["PCT_OCC"]  = np.where(df["TOTAL"] > 0, df["OCC"]  / df["TOTAL"] * 100, 0)

# Integer %s that always sum to 100 (used in the table)
pO_int = np.rint(df["PCT_OCC"].to_numpy()).astype(int)
pO_int = np.clip(pO_int, 0, 100)
pF_int = (100 - pO_int).astype(int)
mask_zero = df["TOTAL"].to_numpy() <= 0
pO_int[mask_zero] = 0
pF_int[mask_zero] = 0

labels = df["Location"].tolist()
T = df["TOTAL"].to_numpy()
F = df["FREE"].to_numpy()
O = df["OCC"].to_numpy()

# ---------------------------------------------------------------------
# 2) COLORS
# ---------------------------------------------------------------------
FREE_COLOR = "#00A878"
OCC_COLOR  = "#3A6EA5"

# ---------------------------------------------------------------------
# 3) PLOT NORMAL STACKED BARS (single axis)
# ---------------------------------------------------------------------
idx = np.arange(len(labels))
w = 0.6

fig, ax = plt.subplots(figsize=(16, 8))

ax.bar(idx, O, width=w, color=OCC_COLOR, edgecolor="white", linewidth=0.5)
ax.bar(idx, F, width=w, bottom=O, color=FREE_COLOR, edgecolor="white", linewidth=0.5)

# Vertical grid lines between bars
ax.set_xticks(idx)
ax.set_xticks(idx + 0.5, minor=True)
ax.grid(which='minor', axis='x', linestyle='-', alpha=0.3, linewidth=0.8)
ax.grid(axis="y", linestyle="--", alpha=0.3)
ax.yaxis.set_major_formatter(StrMethodFormatter("{x:,.0f}"))

# ---------------------------------------------------------------------
# 4) VALUE LABELS INSIDE BARS
# ---------------------------------------------------------------------
for i in range(len(labels)):
    total = float(T[i])
    f = float(F[i])
    o = float(O[i])

    if total <= 0:
        continue

    # OCC value inside
    if o > 0:
        ax.text(i, o / 2, f"{int(o):,}", ha="center", va="center",
                color="black", fontsize=10, fontweight="bold")

    # FREE value inside
    if f > 0:
        ax.text(i, o + f / 2, f"{int(f):,}", ha="center", va="center",
                color="black", fontsize=10, fontweight="bold")

# ---------------------------------------------------------------------
# 5) LEGEND
# ---------------------------------------------------------------------
from matplotlib.patches import Patch

legend_items = [
    Patch(facecolor=OCC_COLOR,  edgecolor='white', label="OCCUPIED"),
    Patch(facecolor=FREE_COLOR, edgecolor='white', label="FREE"),
]

ax.legend(
    handles=legend_items,
    loc="upper right",
    frameon=False,
    fontsize=10
)

# ---------------------------------------------------------------------
# 6) FINAL TOUCHES
# ---------------------------------------------------------------------
fig.suptitle("Stacked OCCUPIED + FREE (Ponderacio)", fontsize=14)
ax.set_xlabel("Location")
ax.set_ylabel("Capacity")
ax.set_xticks(idx)
ax.set_xticklabels(labels, rotation=45, ha="right")

# Reserve bottom margin for the percentage table
plt.tight_layout(rect=[0, 0.14, 1, 1])

# ---------------------------------------------------------------------
# 7) PERCENTAGE TABLE UNDER THE CHART
# ---------------------------------------------------------------------
axTbl = fig.add_axes([0.06, 0.02, 0.88, 0.16])
axTbl.axis("off")

row_labels = ["OCC %", "FREE %"]
cell_text = [
    [f"{v:d}%" for v in pO_int],
    [f"{v:d}%" for v in pF_int],
]

tbl = axTbl.table(
    cellText=cell_text,
    rowLabels=row_labels,
    colLabels=labels,
    loc="center",
    cellLoc="center",
)

tbl.auto_set_font_size(False)
tbl.set_fontsize(10)
tbl.scale(1.0, 1.3)

# ---------------------------------------------------------------------
# 8) SAVE / SHOW
# ---------------------------------------------------------------------
chart_path = in_path.parent / "stacked_percent_palette1(Ponderacio).png"
plt.savefig(chart_path, dpi=220, bbox_inches="tight")
plt.show()


# =============================================================================
# NITROGEN TANKS CHART (unchanged)
# =============================================================================
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patheffects import withStroke
from matplotlib.ticker import StrMethodFormatter

df = nitrogen_grouped.copy()

df.drop(columns=["PONDERACIÓ_TOTAL", "PONDERACIÓ_FREE"], inplace=True, errors="ignore")

required_cols = ["UBICACIO", "CAPACITAT", "LLIURES"]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise ValueError(f"Missing columns required for nitrogen tanks plot: {missing}")

mask_nl = df["UBICACIO"].astype(str).str.match(r"^NL[1-5]$")
nl_df = df.loc[mask_nl, required_cols].copy()

if nl_df.empty:
    raise ValueError("No nitrogen tank rows found (NL1–NL5). Check UBICACIO values or adjust the regex.")

agg = (nl_df.groupby("UBICACIO", as_index=False)
             .agg(TOTAL=("CAPACITAT", "sum"),
                  FREE=("LLIURES", "sum")))
agg["OCC"] = (agg["TOTAL"] - agg["FREE"]).clip(lower=0)

agg["PCT_FREE"] = np.divide(agg["FREE"], agg["TOTAL"], out=np.zeros_like(agg["FREE"], dtype=float), where=agg["TOTAL"]>0) * 100.0
agg["PCT_OCC"]  = np.divide(agg["OCC"],  agg["TOTAL"], out=np.zeros_like(agg["OCC"],  dtype=float), where=agg["TOTAL"]>0) * 100.0

agg.sort_values("TOTAL", ascending=False, inplace=True)

x = agg["UBICACIO"].tolist()
O = agg["OCC"].to_numpy()
F = agg["FREE"].to_numpy()
pO = agg["PCT_OCC"].to_numpy()
pF = agg["PCT_FREE"].to_numpy()

OCC_COLOR  = "#3A6EA5"
FREE_COLOR = "#00A878"

fig, ax = plt.subplots(figsize=(10, 6))

ax.bar(x, O, color=OCC_COLOR, edgecolor="white", linewidth=0.5, label="OCCUPIED")
ax.bar(x, F, bottom=O, color=FREE_COLOR, edgecolor="white", linewidth=0.5, label="FREE")

outline = [withStroke(foreground="black", linewidth=2)]
for i, tank in enumerate(x):
    total = O[i] + F[i]
    if total > 0 and O[i] > 0:
        txt = f"{pO[i]:.0f}%" if pO[i] >= 10 else f"{pO[i]:.1f}%"
        ax.text(i, O[i]/2.0, txt, ha="center", va="center", color="white", fontsize=11, path_effects=outline)
    if total > 0 and F[i] > 0:
        txt = f"{pF[i]:.0f}%" if pF[i] >= 10 else f"{pF[i]:.1f}%"
        ax.text(i, O[i] + F[i]/2.0, txt, ha="center", va="center", color="white", fontsize=11, path_effects=outline)

for i in range(len(x)):
    ax.text(i, O[i]+F[i], f"{O[i]+F[i]:,.0f}", ha="center", va="bottom", fontsize=10)

ax.set_title("Nitrogen Tanks (NL1–NL5): % FREE vs OCCUPIED", fontsize=14)
ax.set_xlabel("Nitrogen tank")
ax.set_ylabel("Capacity")
ax.yaxis.set_major_formatter(StrMethodFormatter("{x:,.0f}"))
ax.grid(axis="y", linestyle="--", alpha=0.35)
ax.legend(loc="upper right")

plt.tight_layout()

out_file = "nitrogen_tanks_stacked.png"
plt.savefig(out_file, dpi=220, bbox_inches="tight")
print(f"Saved figure: {out_file}")

plt.show()


# =============================================================================
# PLOTLY INTERACTIVE CHART — NORMAL (no broken axis, single subplot)
# =============================================================================
import pandas as pd
import numpy as np
import plotly.graph_objects as go

df = grouped.copy()
for c in ["PONDERACIÓ_TOTAL", "PONDERACIÓ_FREE"]:
    if c in df.columns:
        df.drop(columns=c, inplace=True)

df = df[["UBICACIO", "tipus ubic", "CAPACITAT", "LLIURES"]].copy()
df["OCC"] = df["CAPACITAT"] - df["LLIURES"]

df_sub = (df.groupby(["UBICACIO", "tipus ubic"], as_index=False)
            .sum()
            .sort_values(["UBICACIO", "CAPACITAT"], ascending=[True, False])
            .reset_index(drop=True))

# Colour map for each tipus
tipus_unique = df_sub["tipus ubic"].unique().tolist()
palette = [
    "#00A878", "#3A6EA5", "#F28E2B", "#E15759", "#76B7B2",
    "#59A14F", "#EDC948", "#B07AA1", "#FF9DA7", "#9C755F", "#BAB0AC"
]
color_map = {t: palette[i % len(palette)] for i, t in enumerate(tipus_unique)}

# Flatten order for X
x_labels = []
x_tipus  = []
x_ubic   = []
T, F, O  = [], [], []
for ubic, sub in df_sub.groupby("UBICACIO", sort=False):
    for _, r in sub.iterrows():
        x_labels.append(f"{ubic}<br>{r['tipus ubic']}")
        x_tipus.append(r["tipus ubic"])
        x_ubic.append(ubic)
        T.append(float(r["CAPACITAT"]))
        F.append(float(r["LLIURES"]))
        O.append(float(r["OCC"]))

T = np.array(T); F = np.array(F); O = np.array(O)

# Percentages
pF = np.divide(F, T, out=np.zeros_like(F), where=T>0) * 100.0
pO = np.divide(O, T, out=np.zeros_like(O), where=T>0) * 100.0

custom_matrix_all = np.column_stack([x_ubic, x_tipus, T, F, O, pF, pO])

# Build figure with a SINGLE subplot (no broken axis)
fig = go.Figure()

# Prepare arrays by tipus
data_by_tipus = {}
for i, t in enumerate(x_tipus):
    if t not in data_by_tipus:
        data_by_tipus[t] = {"x": [], "O": [], "F": [], "cd": []}
    data_by_tipus[t]["x"].append(x_labels[i])
    data_by_tipus[t]["O"].append(O[i])
    data_by_tipus[t]["F"].append(F[i])
    data_by_tipus[t]["cd"].append(custom_matrix_all[i, :])

# Hover templates
hover_occ = (
    "<b>%{customdata[0]}</b> — %{customdata[1]}<br>"
    "Series: <b>OCCUPIED</b><br>"
    "TOTAL: %{customdata[2]:,.0f}<br>"
    "OCC: %{y:,.0f} (<b>%{customdata[6]:.1f}%</b>)<br>"
    "FREE: %{customdata[3]:,.0f} (%{customdata[5]:.1f}%)<br>"
    "<extra></extra>"
)
hover_free = (
    "<b>%{customdata[0]}</b> — %{customdata[1]}<br>"
    "Series: <b>FREE</b><br>"
    "TOTAL: %{customdata[2]:,.0f}<br>"
    "FREE: %{y:,.0f} (<b>%{customdata[5]:.1f}%</b>)<br>"
    "OCC: %{customdata[4]:,.0f} (%{customdata[6]:.1f}%)<br>"
    "<extra></extra>"
)

for t, dct in data_by_tipus.items():
    color = color_map.get(t, "#888888")
    x = dct["x"]
    occ_y = dct["O"]
    free_y = dct["F"]
    cd = np.array(dct["cd"])

    # OCC trace
    fig.add_trace(
        go.Bar(
            name=f"OCC — {t}",
            x=x, y=occ_y,
            marker_color=color,
            opacity=0.65,
            customdata=cd,
            hovertemplate=hover_occ,
        )
    )
    # FREE trace
    fig.add_trace(
        go.Bar(
            name=f"FREE — {t}",
            x=x, y=free_y,
            marker_color=color,
            opacity=1.0,
            customdata=cd,
            hovertemplate=hover_free,
        )
    )

# Stacking mode
fig.update_layout(barmode="stack")

# X ticks
fig.update_xaxes(tickangle=90)
fig.update_yaxes(title="Capacity")

# Group shading (shapes spanning full y)
starts = []
ends   = []
pos = 0
for ubic, sub in df_sub.groupby("UBICACIO", sort=False):
    n = len(sub)
    starts.append(pos)
    ends.append(pos + n - 1)
    pos += n

for j, (s, e) in enumerate(zip(starts, ends)):
    x0 = x_labels[s]
    x1 = x_labels[e]
    shade = "rgba(243,246,250,0.9)" if j % 2 == 0 else "rgba(255,255,255,1.0)"
    fig.add_shape(
        type="rect",
        xref="x", yref="paper",
        x0=x0, x1=x1,
        y0=0, y1=1,
        fillcolor=shade,
        line=dict(width=0),
        layer="below"
    )

fig.update_layout(
    title=dict(
        text="NO PONDERACIO",
        x=0.5
    ),
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0.0),
    hovermode="x unified",
    margin=dict(l=40, r=40, t=60, b=80),
    height=700
)

fig.write_html("No_PON_subcategories_interactive.html", include_plotlyjs="cdn", auto_open=True)
print("Saved: No_PON_subcategories_interactive.html")


# =============================================================================
# PDF REPORT (unchanged)
# =============================================================================
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
import datetime
import pandas as pd
from pathlib import Path


def _sanitize_text(s: str) -> str:
    if not isinstance(s, str):
        return s
    repl = {
        "\u2013": "-", "\u2014": "-", "\u2022": "•",
        "\u00A0": " ", "\u200b": "", "\u2060": "",
        "\uf0b7": "•"
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def _fmt_num(v, decimals=1):
    try:
        if pd.isna(v):
            return ""
        return f"{float(v):,.{decimals}f}"
    except:
        return str(v)


def _df_round_copy(df: pd.DataFrame,
                   weighted_cols=("PONDERACIÓ_TOTAL","PONDERACIÓ_FREE"),
                   int_cols=("CAPACITAT","LLIURES")) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if c in weighted_cols and pd.api.types.is_numeric_dtype(out[c]):
            out[c] = out[c].round(1)
        elif c in int_cols and pd.api.types.is_numeric_dtype(out[c]):
            out[c] = out[c].round(0).astype("Int64")
    return out


def _make_table(df: pd.DataFrame, col_widths=None, zebra=True,
                right_align_numbers=True, header_bg=colors.whitesmoke,
                add_totals_row=False, conditional_free_shading=False,
                pct_col=None, low_thresh=10, high_thresh=70):

    _df = df.copy()

    if add_totals_row:
        total_dict = {}
        for c in _df.columns:
            if pd.api.types.is_numeric_dtype(_df[c]):
                total_value = _df[c].sum(skipna=True)
                if "PONDERACIÓ" in c.upper():
                    total_dict[c] = round(total_value, 1)
                elif c.upper() in ("CAPACITAT", "LLIURES"):
                    total_dict[c] = int(round(total_value))
                else:
                    total_dict[c] = round(total_value, 1)
            else:
                total_dict[c] = "TOTAL"
        _df.loc[len(_df)] = total_dict

    data = [list(_df.columns)] + _df.fillna("").values.tolist()
    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    style = [
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("BACKGROUND", (0,0), (-1,0), header_bg),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("LINEABOVE", (0,0), (-1,0), 0.75, colors.grey),
        ("LINEBELOW", (0,0), (-1,0), 0.75, colors.grey),
        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,-1), 9),
        ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey)
    ]

    if right_align_numbers:
        for j, col in enumerate(_df.columns):
            if pd.api.types.is_numeric_dtype(_df[col]):
                style.append(("ALIGN", (j,1), (j,-1), "RIGHT"))

    if zebra:
        for r in range(1, len(data)):
            if r % 2 == 1:
                style.append(("BACKGROUND", (0,r), (-1,r), colors.Color(0.97, 0.97, 0.97)))

    if conditional_free_shading and pct_col in _df.columns:
        pct_idx = _df.columns.get_loc(pct_col)
        for r in range(1, len(data)):
            try:
                v = float(_df.iloc[r-1, pct_idx])
            except:
                v = None
            if v is None:
                continue
            if v <= low_thresh:
                style.append(("BACKGROUND", (0,r), (-1,r), colors.Color(1.0, 0.92, 0.92)))
            elif v >= high_thresh:
                style.append(("BACKGROUND", (0,r), (-1,r), colors.Color(0.92, 1.0, 0.92)))

    tbl.setStyle(TableStyle(style))
    return tbl


def _get_cols_by_alias(df: pd.DataFrame):
    colmap = {c.upper(): c for c in df.columns}

    loc = None
    for key in ("UBICACIO", "LOCATION"):
        if key in colmap:
            loc = colmap[key]; break

    ptotal = None
    for key in ("PONDERACIÓ_TOTAL","PONDERACIO_TOTAL","SUM OF PONDERACIÓ_TOTAL","SUM OF PONDERACIO_TOTAL"):
        if key in colmap:
            ptotal = colmap[key]; break

    pfree = None
    for key in ("PONDERACIÓ_FREE","PONDERACIO_FREE","SUM OF PONDERACIÓ_FREE","SUM OF PONDERACIO_FREE"):
        if key in colmap:
            pfree = colmap[key]; break

    return loc, ptotal, pfree


def build_pdf_report(
    total_space,
    total_free_space,
    growth_df,
    totals_by_type_df,
    by_location_totals,
    chart_path,
    out_pdf,
    top_n=3,
    report_title="FREEZERS Summary Report (Weighted Ponderació)"
):

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleLarge", fontSize=20, leading=24, spaceAfter=14))
    styles.add(ParagraphStyle(name="H2", fontSize=14, spaceBefore=12, spaceAfter=6))
    styles.add(ParagraphStyle(name="Body", fontSize=11, leading=15))
    styles.add(ParagraphStyle(name="Small", fontSize=9, leading=12))

    doc = SimpleDocTemplate(
        str(out_pdf), pagesize=A4,
        leftMargin=1.7*cm, rightMargin=1.7*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    loc_col, ptotal_col, pfree_col = _get_cols_by_alias(by_location_totals)

    locdf = by_location_totals[[loc_col, ptotal_col, pfree_col]].copy()
    locdf.rename(columns={
        loc_col: "Location",
        ptotal_col: "W_TOTAL",
        pfree_col: "W_FREE"
    }, inplace=True)

    locdf["FREE_%"] = (
        locdf["W_FREE"] / locdf["W_TOTAL"].replace({0: pd.NA}) * 100
    ).round(1).fillna(0)

    top_free = locdf.sort_values("W_FREE", ascending=False).head(top_n)

    saturated = locdf[locdf["W_TOTAL"] > 0].copy()
    most_saturated = saturated.sort_values("FREE_%", ascending=True).head(3)

    free_pct = (total_free_space / total_space * 100) if total_space else 0
    occ_pct = 100 - free_pct

    story = []

    story.append(Paragraph(_sanitize_text(report_title), styles["TitleLarge"]))
    story.append(Paragraph(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Body"]))
    story.append(Spacer(1, 0.4*cm))

    intro = (
        "This report summarizes the current storage distribution across ultra-low freezers using "
        "weighted capacity (Ponderacio) to standardize space usage across different box types. "
        "It highlights storage availability, growth potential, and areas with high strategic value for expansion."
    )
    story.append(Paragraph(_sanitize_text(intro), styles["Body"]))
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("Summary", styles["H2"]))
    story.append(Paragraph(f"<b>Total Space (weighted):</b> {_fmt_num(total_space,1)}", styles["Body"]))
    story.append(Paragraph(f"<b>Total Free Space (weighted):</b> {_fmt_num(total_free_space,1)}", styles["Body"]))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("Key Information", styles["H2"]))
    story.append(Paragraph(f"- <b>Total Occupancy:</b> {_fmt_num(occ_pct,1)}%", styles["Body"]))
    story.append(Paragraph(f"- <b>Total Free Percentage:</b> {_fmt_num(free_pct,1)}%", styles["Body"]))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph(
        _sanitize_text(
            "These percentages represent the weighted storage capacity, meaning each freezer box type contributes "
            "proportionally based on its volume scaling factor (0.5 mL, 1.4 mL, 9x9, Gradetes). "
            "Higher-capacity box types contribute more weight toward the overall totals."
        ),
        styles["Body"]
    ))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Key Findings", styles["H2"]))
    findings = [
        f"- Total weighted storage: {_fmt_num(total_space,1)}",
        f"- Free weighted storage: {_fmt_num(total_free_space,1)}",
        f"- System-wide free capacity: {_fmt_num(free_pct,1)}%",
        f"- Highest potential locations: {', '.join(top_free['Location'])}",
        f"- Most saturated locations: {', '.join(most_saturated['Location'])}",
    ]
    for line in findings:
        story.append(Paragraph(_sanitize_text(line), styles["Body"]))

    story.append(PageBreak())

    story.append(Paragraph("Stacked OCCUPIED + FREE (Weighted Ponderacio)", styles["H2"]))

    if chart_path and Path(chart_path).exists():
        img = Image(str(chart_path))
        img._restrictSize(15.5*cm, 10.5*cm)
        story.append(img)
    else:
        story.append(Paragraph("<b>Chart not found.</b>", styles["Body"]))

    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph(
        _sanitize_text(
            "The distribution shows a clear separation between high-availability locations and more saturated ones. "
            "Freezers with large weighted FREE values indicate strong capacity for expansion and redistribution. "
            "Locations with very low FREE% may require closer operational monitoring to prevent congestion."
        ),
        styles["Body"]
    ))
    story.append(Spacer(1, 0.3*cm))

    story.append(
        Paragraph(
            _sanitize_text("Top free-capacity locations (weighted): " + ", ".join(top_free["Location"])),
            styles["Body"]
        )
    )

    story.append(PageBreak())

    story.append(Paragraph("Maximal Potential Growth by Box Type", styles["H2"]))
    story.append(_make_table(_df_round_copy(growth_df)))
    story.append(Spacer(1, 0.3*cm))

    mpg_text = (
        "Interpretation:<br/>"
        "This table estimates how much additional storage capacity can be achieved if the currently "
        "available weighted free space is filled exclusively with one box type at a time. "
        "Larger-volume box formats (such as Gradetes and 9x9) provide a lower count of additional boxes "
        "for the same amount of weighted free space, while smaller box types (such as Willmut 0.5 mL) yield "
        "higher potential growth values. These values help determine the most efficient box formats to "
        "prioritize when planning future storage expansion or allocating resources for new projects."
    )
    story.append(Paragraph(_sanitize_text(mpg_text), styles["Body"]))
    story.append(Spacer(1, 0.8*cm))

    tbt = totals_by_type_df.copy()

    if all(c in tbt.columns for c in ["tipus ubic","CAPACITAT","LLIURES","PONDERACIÓ_TOTAL","PONDERACIÓ_FREE"]):
        tbt = tbt[["tipus ubic","CAPACITAT","LLIURES","PONDERACIÓ_TOTAL","PONDERACIÓ_FREE"]]

    if "PONDERACIÓ_TOTAL" in tbt.columns and "PONDERACIÓ_FREE" in tbt.columns:
        tbt["_FREE_%"] = (
            tbt["PONDERACIÓ_FREE"] / tbt["PONDERACIÓ_TOTAL"].replace({0: pd.NA}) * 100
        ).round(1).fillna(0)
    else:
        tbt["_FREE_%"] = 0

    tbt_display = _df_round_copy(
        tbt.drop(columns=["_FREE_%"]),
        weighted_cols=("PONDERACIÓ_TOTAL","PONDERACIÓ_FREE"),
        int_cols=("CAPACITAT","LLIURES")
    )

    story.append(Paragraph("Totals by Type", styles["H2"]))
    story.append(
        _make_table(
            tbt_display,
            zebra=True,
            right_align_numbers=True,
            add_totals_row=True,
            conditional_free_shading=True,
            pct_col="_FREE_%"
        )
    )
    story.append(Spacer(1, 0.4*cm))

    totals_text = (
        "Interpretation:<br/>"
        "This table summarizes the total raw capacity, current free space, and weighted (Ponderacio) storage values "
        "for each box type across all freezers. The weighted columns reflect how much effective freezer volume each "
        "box type occupies, enabling comparison across formats of different physical sizes."
    )
    story.append(Paragraph(_sanitize_text(totals_text), styles["Body"]))

    doc.build(story)


# === Build the PDF Report ===
chart_path = Path("stacked_percent_palette1(Ponderacio).png")
out_pdf = in_path.parent / "FREEZERS_report.pdf"

build_pdf_report(
    total_space=total_space,
    total_free_space=total_free_space,
    growth_df=growth_df,
    totals_by_type_df=by_type,
    by_location_totals=by_location_totals,
    chart_path=chart_path,
    out_pdf=out_pdf,
    top_n=3
)

print("PDF report created:", out_pdf)

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# Config
# =========================
# >>> UPDATE THESE TWO LINES BEFORE EACH RUN <<<
# Change the date (YYYYMMDD) to match your exported file.
# Example: "GENERAL_STORAGE_20260301.xlsx"
INPUT_FILE  = "GENERAL_STORAGE_20260205.xlsx"
OUTPUT_FILE = "SERVICES_SUMMARY.xlsx"


# Exact label in UBIC_TIPUS for cryobox
CRYOBOX_LABEL = "CRIOBOX Nitrogen 10x10"

# =========================
# Safety: verify input exists
# =========================
if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(
        f"Input file not found: {INPUT_FILE}\n"
        f"Working directory: {os.getcwd()}\n"
        f"Please place the Excel file in this folder or update INPUT_FILE."
    )

# =========================
# Load
# =========================
df = pd.read_excel(INPUT_FILE)

# =========================
# Step 1 — Normalize UBIC_TIPUS (mapping the four families)
# =========================
replace_map = {
    # Wilmut 0,5
    "Calaix ample per microplaques (-80ºC)": "Wilmut 0,5",
    "Calaix estret per microplaques (-80ºC)": "Wilmut 0,5",
    "Calaix SOST microplaques (-80ºC)": "Wilmut 0,5",

    # Wilmut 1.4 mL
    "Calaix per microplaques grans 1,4ml (-80ºC)": "Wilmut 1.4 mL",
    "Calaix SOST microplaques grans (-80ºC)": "Wilmut 1.4 mL",

    # 9x9
    "Calaix CAIXES ALTES 9x9 (-80ºC)": "9x9",
    "Calaix per CAIXES 9x9 (-80ºC)": "9x9",
    "Calaix SOST CAIXES 9X9 (-80ºC)": "9x9",

    # Gradetes
    "Estant per Gradetes 10x4": "Gradetes",
    "Gradetes tubs 4 posiciones": "Gradetes",
    "Gradetes tubs 5x12": "Gradetes",
}
df["UBIC_TIPUS"] = df["UBIC_TIPUS"].replace(replace_map)

# =========================
# Step 2 — Keep required columns
# =========================
df = df[["SERVEI", "UBIC_TIPUS"]]

# =========================
# Step 3 — Filter SERVEI starting with A0 or A3
# =========================
df = df[df["SERVEI"].astype(str).str.startswith(("A0", "A3"))]

# =========================
# Step 4 — Build Data sheet (pivot + ponderades + totals)
# =========================
pivot = (
    df.pivot_table(index="SERVEI", columns="UBIC_TIPUS", aggfunc="size", fill_value=0)
      .reset_index()
)

required_base_cols = ["SERVEI", "Wilmut 0,5", "Wilmut 1.4 mL", "9x9", "Gradetes"]
for col in required_base_cols:
    if col not in pivot.columns:
        pivot[col] = 0
pivot = pivot[required_base_cols]

# Add PONDERADES
pivot["PONDERADES Wilmut 0,5"]   = pivot["Wilmut 0,5"]    * 1
pivot["PONDERADES Wilmut 1.4mL"] = pivot["Wilmut 1.4 mL"] * 2
pivot["PONDERADES 9x9"]          = pivot["9x9"]           * 2.55
pivot["PONDERADES Gradetes"]     = pivot["Gradetes"]      * 10.2

# Arrange columns and add TOTAL PONDERACIÓ
ordered_cols = [
    "SERVEI",
    "Wilmut 0,5", "PONDERADES Wilmut 0,5",
    "Wilmut 1.4 mL", "PONDERADES Wilmut 1.4mL",
    "9x9", "PONDERADES 9x9",
    "Gradetes", "PONDERADES Gradetes",
]
pivot = pivot[ordered_cols]
pivot["TOTAL PONDERACIÓ"] = (
    pivot["PONDERADES Wilmut 0,5"]
    + pivot["PONDERADES Wilmut 1.4mL"]
    + pivot["PONDERADES 9x9"]
    + pivot["PONDERADES Gradetes"]
)

# Grand TOTAL row for Data
grand_totals_data = {"SERVEI": "TOTAL"}
for col in pivot.columns:
    if col != "SERVEI":
        grand_totals_data[col] = pivot[col].sum()
pivot = pd.concat([pivot, pd.DataFrame([grand_totals_data])], ignore_index=True)

# =========================
# Step 5 — Build Cryobox sheet (ONLY CRIOBOX Nitrogen 10x10; NO PONDERADES column)
# =========================
df_cryo = df[df["UBIC_TIPUS"] == CRYOBOX_LABEL].copy()

if not df_cryo.empty:
    cryobox_summary = (
        df_cryo.groupby("SERVEI")
               .size()
               .reset_index(name=CRYOBOX_LABEL)
               .sort_values("SERVEI")
    )
else:
    cryobox_summary = pd.DataFrame(columns=["SERVEI", CRYOBOX_LABEL])

# Add total row
total_cryo = cryobox_summary[CRYOBOX_LABEL].sum() if not cryobox_summary.empty else 0
cryobox_summary = pd.concat(
    [cryobox_summary, pd.DataFrame([{"SERVEI": "TOTAL", CRYOBOX_LABEL: total_cryo}])],
    ignore_index=True
)
cryobox_summary = cryobox_summary[["SERVEI", CRYOBOX_LABEL]]

# =========================
# Step 6 — Write BOTH sheets in one go (no append)
# =========================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    pivot.to_excel(writer, sheet_name="Data", index=False)
    cryobox_summary.to_excel(writer, sheet_name="Cryobox", index=False)

# =========================
# Step 7 — Formatting (thin borders, header color, zebra, numbers)
# =========================
wb = load_workbook(OUTPUT_FILE)

HEADER_FILL = "1F4E78"  # dark blue
HEADER_FONT = "FFFFFF"  # white
ZEBRA_FILL  = "F2F2F2"  # light gray

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_sheet(ws, number_map):
    # Freeze header + AutoFilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Body borders + zebra striping
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = thin_border
        if r % 2 == 0:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=ZEBRA_FILL)

    # Number formats per column letter
    for col_letter, kind in number_map.items():
        for r in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{r}"]
            if kind == "int":
                cell.number_format = "#,##0"
            elif kind == "float":
                cell.number_format = "#,##0.00"

    # Auto-fit widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

# Format Data (A..J expected: raw B,D,F,H ints; ponderades & total C,E,G,I,J floats)
ws_data = wb["Data"]
data_number_map = {"B": "int", "D": "int", "F": "int", "H": "int",
                   "C": "float", "E": "float", "G": "float", "I": "float", "J": "float"}
format_sheet(ws_data, data_number_map)

# Format Cryobox (A: text, B: integer counts)
ws_cryo = wb["Cryobox"]
cryo_number_map = {"B": "int"}
format_sheet(ws_cryo, cryo_number_map)

wb.save(OUTPUT_FILE)
print(f" ✔  Done! {OUTPUT_FILE} SAVED. Sheets created: Data, Cryobox")
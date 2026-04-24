# -*- coding: utf-8 -*-
"""
Created on Thu Apr 23 10:20:43 2026

@author: LCHIBUIKE
"""

import pandas as pd
import glob
import os
import inspect
from functools import reduce
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# -------------------------------------------------------------
# USER INPUT: YEARS TO ANALYZE
# -------------------------------------------------------------
YEARS_TO_ANALYZE = [2024, 2025, 2026]   # ✅ EDIT THIS LIST AS NEEDED

# -------------------------------------------------------------
# FUNCTIONS
# -------------------------------------------------------------
def has_year_date(cell, year):
    if pd.isna(cell):
        return False
    for part in str(cell).split(','):
        dt = pd.to_datetime(part.strip(), dayfirst=True, errors="coerce")
        if pd.notna(dt) and dt.year == year:
            return True
    return False


def autofit(ws):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = (
            max(len(str(c.value)) for c in col if c.value) + 2
        )

# -------------------------------------------------------------
# STEP 1 — Script folder
# -------------------------------------------------------------
try:
    script_folder = os.path.dirname(os.path.abspath(__file__))
except NameError:
    script_folder = os.getcwd()

# -------------------------------------------------------------
# STEP 2 — Detect sample Excel files
# -------------------------------------------------------------
excel_files = [
    f for f in glob.glob(os.path.join(script_folder, "*.xlsx"))
    if "_samples_" in os.path.basename(f).lower()
]

# Column names
delivery_col = "Delivery date"
alt_code_col = "Request alternative code"
case_col = "[Case] NorayBanks case code"

# -------------------------------------------------------------
# Containers
# -------------------------------------------------------------
results = []
year_rows = {}   # year → list of filtered DataFrames

# -------------------------------------------------------------
# STEP 3 — Process each file (USER‑DEFINED YEARS)
# -------------------------------------------------------------
for file in excel_files:
    file_name = os.path.basename(file)
    print(f"🔄 Processing file: {file_name}")

    df = pd.read_excel(file, header=1)
    df.columns = df.columns.str.strip()
    df[delivery_col] = df[delivery_col].astype(str)

    cohort = file_name.split("_")[0]
    row = {"Cohort": cohort}

    for year in YEARS_TO_ANALYZE:
        df_year = df[df[delivery_col].apply(lambda x: has_year_date(x, year))]
        df_year_filtered = df_year[
            df_year[alt_code_col]
            .astype(str)
            .str.upper()
            .str.strip()
            .str.startswith("A")
        ]

        row[f"procurement_{year}"] = df_year_filtered.shape[0]
        row[f"cases_{year}"] = df_year_filtered[case_col].nunique()

        if not df_year_filtered.empty:
            year_rows.setdefault(year, []).append(df_year_filtered)

    results.append(row)

results_df = pd.DataFrame(results)

# -------------------------------------------------------------
# STEP 4 — % Change (First year → Second year, if provided)
# -------------------------------------------------------------

# -------------------------------------------------------------
# STEP 5 — Write Excel output
# -------------------------------------------------------------
output_file = os.path.join(script_folder, "procurement_summary.xlsx")

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    # =========================
    # SUMMARY SHEET
    # =========================
    total_row = {"Cohort": "TOTAL"}
    for col in results_df.columns:
        if col != "Cohort":
            total_row[col] = results_df[col].sum(skipna=True)

    summary_df = pd.concat([results_df, pd.DataFrame([total_row])], ignore_index=True)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    ws = writer.book["Summary"]
    for cell in ws[1]:
        cell.font = Font(bold=True)

    table = Table(
        displayName="SummaryTable",
        ref=f"A1:{chr(64 + ws.max_column)}{ws.max_row}"
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True
    )
    ws.add_table(table)
    autofit(ws)

    # =========================
    # SAMPLES BY TYPE — PER YEAR
    # =========================
    for year, frames in year_rows.items():
        combined = pd.concat(frames, ignore_index=True)

        if "Sample type" not in combined.columns:
            continue

        grouped = (
            combined.groupby("Sample type")
            .size()
            .reset_index(name="Number of Samples")
        )

        grouped = pd.concat([
            grouped,
            pd.DataFrame([{
                "Sample type": "TOTAL",
                "Number of Samples": grouped["Number of Samples"].sum()
            }])
        ], ignore_index=True)

        sheet_name = f"Samples_by_Type_{year}"
        grouped.to_excel(writer, sheet_name=sheet_name, index=False)

        ws_y = writer.book[sheet_name]
        for cell in ws_y[1]:
            cell.font = Font(bold=True)

        table = Table(
            displayName=f"SampleType{year}",
            ref=f"A1:B{ws_y.max_row}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )
        ws_y.add_table(table)
        autofit(ws_y)

    # =========================
    # PIVOTED SAMPLE TYPE COMPARISON
    # =========================
    pivot_frames = []

    for year, frames in year_rows.items():
        combined = pd.concat(frames, ignore_index=True)
        if "Sample type" in combined.columns:
            pivot_frames.append(
                combined.groupby("Sample type")
                .size()
                .reset_index(name=str(year))
            )

    if pivot_frames:
        pivot_final = reduce(
            lambda l, r: pd.merge(l, r, on="Sample type", how="outer"),
            pivot_frames
        ).fillna(0)

        pivot_final.to_excel(
            writer,
            sheet_name="Samples_Type_Comparison",
            index=False
        )

        ws_p = writer.book["Samples_Type_Comparison"]
        for cell in ws_p[1]:
            cell.font = Font(bold=True)

        table = Table(
            displayName="SampleTypePivot",
            ref=f"A1:{chr(64 + ws_p.max_column)}{ws_p.max_row}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )
        ws_p.add_table(table)
        autofit(ws_p)

# -------------------------------------------------------------
# STEP 6 — Save script copy
# -------------------------------------------------------------
try:
    this_script = inspect.getsource(inspect.getmodule(inspect.currentframe()))
    save_path = os.path.join(script_folder, "procurement_script.py")
    with open(save_path, "w", encoding="utf-8") as f:
        f.write(this_script)
except Exception:
    save_path = "Script saving skipped."

# -------------------------------------------------------------
# STEP 7 — Completion message
# -------------------------------------------------------------
print("\n--------------------------------------------")
print("✅ Processing completed successfully!")
print(f"✅ Total files processed: {len(excel_files)}")
print(f"✅ Years analyzed: {YEARS_TO_ANALYZE}")
print(f"✅ Results saved to: {output_file}")
print(f"✅ Script saved to: {save_path}")
print("--------------------------------------------\n")